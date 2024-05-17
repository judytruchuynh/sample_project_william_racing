import json
from kafka import KafkaConsumer
from collections import defaultdict, Counter
import time
from openpyxl import Workbook, load_workbook

# Define the Kafka topic
TOPIC_NAME = 'f1_telemetry'

# Create a Kafka consumer
consumer = KafkaConsumer(
    TOPIC_NAME,
    bootstrap_servers='localhost:9092',
    value_deserializer=lambda v: json.loads(v.decode('utf-8')),
    auto_offset_reset='earliest',
    enable_auto_commit=True,
    group_id='f1_group'
)

max_speed = 0
total_throttle = 0
total_records = 0
max_rpm = 0
total_lap_time = 0
gear_counter = Counter()

# Create or load the Excel workbook
try:
    wb = load_workbook('output.xlsx')
except FileNotFoundError:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title='Telemetry Data')
    # Add schema
    ws.append([
        'Speed (km/h)',
        'RPM',
        'Throttle (%)',
        'Lap Time (s)',
        'Gear',
        'Max Speed (km/h)',
        'Max RPM',
        'Average Throttle (%)',
        'Average Lap Time (s)',
        'Most Frequent Gear'
    ])
    wb.save('output.xlsx')
ws = wb.active

print("Starting consumer...")

while True:
    for message in consumer:
        telemetry_data = message.value

        # Update max speed
        if telemetry_data['speed'] > max_speed:
            max_speed = telemetry_data['speed']

        # Update max RPM
        if telemetry_data['rpm'] > max_rpm:
            max_rpm = telemetry_data['rpm']

        # Calculate total throttle
        total_throttle += telemetry_data['throttle']

        # Calculate total lap time
        total_lap_time += telemetry_data['lap_time']

        # Count the frequency of each gear
        gear_counter[telemetry_data['gear']] += 1

        total_records += 1

        # Calculate average throttle
        average_throttle = total_throttle / total_records

        # Calculate average lap time
        average_lap_time = total_lap_time / total_records

        # Determine the most frequent gear
        most_frequent_gear = gear_counter.most_common(1)[0][0]

        print(f"Telemetry Data: {telemetry_data}")
        print(f"Max Speed: {max_speed:.2f} km/h")
        print(f"Max RPM: {max_rpm:.2f}")
        print(f"Average Throttle: {average_throttle:.2f}%")
        print(f"Average Lap Time: {average_lap_time:.2f} seconds")
        print(f"Most Frequent Gear: {most_frequent_gear}")

        # Write data to Excel
        ws.append([
            telemetry_data['speed'],
            telemetry_data['rpm'],
            telemetry_data['throttle'],
            telemetry_data['lap_time'],
            telemetry_data['gear'],
            max_speed,
            max_rpm,
            average_throttle,
            average_lap_time,
            most_frequent_gear
        ])

        # Save the workbook
        wb.save('output.xlsx')

        # Additional analysis can be added here as needed

    # Wait for 15 seconds before processing next set of records
    time.sleep(15)
